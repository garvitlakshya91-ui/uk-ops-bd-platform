"""
Scraper for Local Authority Housing Statistics (LAHS).

LAHS is published annually by DLUHC and provides council-level statistics on:
- Social housing stock counts (council-owned, RP-owned)
- Right to Buy sales
- Homelessness figures
- Housing waiting list sizes
- New build completions
- Affordable housing supply

This data enriches our council records and helps score BD opportunities
by identifying councils with the most housing stock and activity.

Data source: https://www.gov.uk/government/collections/local-authority-housing-data

The dataset is published as an XLSX/ODS spreadsheet. We download the
latest version and extract key metrics per local authority.
"""

from __future__ import annotations

import io
import re
from typing import Any

import httpx
import structlog

logger = structlog.get_logger(__name__)

# GOV.UK page listing LAHS releases
LAHS_INDEX_URL = "https://www.gov.uk/government/statistical-data-sets/local-authority-housing-statistics-data-returns-for-2022-to-2023"

# Direct download URL patterns (updated annually)
# Typical: https://assets.publishing.service.gov.uk/media/...LAHS_Open_Data_2022-23.xlsx
LAHS_FALLBACK_URL = (
    "https://assets.publishing.service.gov.uk/media/"
    "65e8f5e6cf7eb1000e7b252a/LAHS_Open_Data_2022-23.xlsx"
)


class LAHSScraper:
    """Scraper for Local Authority Housing Statistics.

    Downloads the LAHS open data XLSX and extracts key metrics per
    local authority code / name.

    Usage::

        scraper = LAHSScraper()
        data = await scraper.fetch_lahs_data()
        # data = [{"la_name": "...", "la_code": "...", "total_stock": ..., ...}, ...]
    """

    def __init__(self, download_url: str | None = None) -> None:
        self._download_url = download_url or LAHS_FALLBACK_URL

    async def fetch_lahs_data(self) -> list[dict[str, Any]]:
        """Download and parse the LAHS open data XLSX.

        Returns a list of dicts, one per local authority, with keys:
        - la_name: Local authority name
        - la_code: ONS local authority code (e.g. E09000001)
        - total_stock: Total local authority housing stock
        - rp_stock: Registered Provider stock in the area
        - waiting_list: Housing waiting list size
        - new_builds: New build completions in the period
        - affordable_supply: Affordable housing supply (gross)
        - rtb_sales: Right to Buy sales in the period
        """
        try:
            import openpyxl
        except ImportError:
            logger.error("lahs_openpyxl_not_installed")
            raise ImportError(
                "openpyxl is required for LAHS parsing. "
                "Install it with: pip install openpyxl"
            )

        logger.info("lahs_download_start", url=self._download_url)

        async with httpx.AsyncClient(timeout=httpx.Timeout(120.0, connect=30.0)) as client:
            resp = await client.get(self._download_url, follow_redirects=True)
            resp.raise_for_status()

        logger.info("lahs_download_complete", bytes=len(resp.content))

        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)

        results: list[dict[str, Any]] = []

        # Try to find the main data sheet — LAHS naming varies by year
        target_sheet = None
        for name in wb.sheetnames:
            lower = name.lower()
            if "stock" in lower or "section 1" in lower or "data" in lower:
                target_sheet = name
                break

        if not target_sheet:
            # Fall back to first sheet
            target_sheet = wb.sheetnames[0]

        ws = wb[target_sheet]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            logger.warning("lahs_empty_sheet", sheet=target_sheet)
            return []

        # Find header row (contains "local authority" or "la name" or "ONS code")
        header_idx = 0
        for i, row in enumerate(rows[:20]):
            row_str = " ".join(str(c or "").lower() for c in row)
            if "local authority" in row_str or "la name" in row_str or "ons code" in row_str:
                header_idx = i
                break

        headers = [str(c or "").strip().lower() for c in rows[header_idx]]

        # Map column indices
        def _find_col(*patterns: str) -> int | None:
            for p in patterns:
                for i, h in enumerate(headers):
                    if p in h:
                        return i
            return None

        col_name = _find_col("local authority name", "la name", "local authority")
        col_code = _find_col("ons code", "la code", "code")
        col_stock = _find_col("total stock", "total dwelling", "la stock")
        col_rp = _find_col("rp stock", "registered provider", "ha stock")
        col_waiting = _find_col("waiting list", "housing register")
        col_builds = _find_col("new build", "completions")
        col_affordable = _find_col("affordable", "gross supply")
        col_rtb = _find_col("right to buy", "rtb")

        if col_name is None:
            logger.warning("lahs_header_not_found", headers=headers[:10])
            return []

        # Parse data rows
        for row in rows[header_idx + 1:]:
            name = str(row[col_name] or "").strip() if col_name is not None else ""
            if not name or name.lower() in ("total", "england", "all", ""):
                continue

            # Skip summary/subtotal rows
            if any(kw in name.lower() for kw in ["region", "total", "of which"]):
                continue

            entry: dict[str, Any] = {
                "la_name": name,
                "la_code": str(row[col_code] or "").strip() if col_code is not None else "",
            }

            for field, col in [
                ("total_stock", col_stock),
                ("rp_stock", col_rp),
                ("waiting_list", col_waiting),
                ("new_builds", col_builds),
                ("affordable_supply", col_affordable),
                ("rtb_sales", col_rtb),
            ]:
                if col is not None and col < len(row):
                    try:
                        val = row[col]
                        entry[field] = int(val) if val is not None else None
                    except (ValueError, TypeError):
                        entry[field] = None
                else:
                    entry[field] = None

            results.append(entry)

        logger.info("lahs_parse_complete", authorities=len(results))
        wb.close()
        return results
