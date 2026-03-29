"""RSH Registered Providers list and Statistical Data Return (SDR) scraper.

Two data sources from the Regulator of Social Housing:

1. **Registered Providers list** (monthly CSV/XLSX)
   Lists every UK registered provider (RP) of social housing with their
   RSH registration number, name, type (PRP/LRP), status, address, and
   stock size. ~1,600 providers.
   Source: https://www.gov.uk/guidance/find-a-registered-provider

2. **Statistical Data Return (SDR)** (annual XLSX)
   Self-reported stock data: units *owned* vs units *managed*. A provider
   that manages more than it owns is operating on behalf of another owner —
   this directly informs the operator/owner split in ExistingScheme.
   Source: https://www.gov.uk/government/collections/statistical-data-return

Usage::

    scraper = RSHRegisteredProvidersScraper()
    providers = await scraper.fetch_registered_providers()
    sdr = await scraper.fetch_sdr_stock()
"""

from __future__ import annotations

import io
import re
from typing import Any

import httpx
import structlog
from bs4 import BeautifulSoup

logger = structlog.get_logger(__name__)

_GOV_BASE = "https://www.gov.uk"

# GOV.UK publication pages to check for the latest file
_RP_LIST_URL = "https://www.gov.uk/guidance/find-a-registered-provider"
_RP_PUBLICATION_URL = (
    "https://www.gov.uk/government/publications/"
    "registered-providers-of-social-housing-in-england"
)
_SDR_COLLECTION_URL = (
    "https://www.gov.uk/government/collections/statistical-data-return"
)
# The SDR is now published as "Registered provider social housing stock and rents"
# These URLs are checked in order from most-recent to older fallbacks
_SDR_PUBLICATION_URLS = [
    "https://www.gov.uk/government/statistics/registered-provider-social-housing-stock-and-rents-in-england-2024-to-2025",
    "https://www.gov.uk/government/statistics/private-registered-provider-social-housing-stock-and-rents-in-england-2024-to-2025",
    "https://www.gov.uk/government/statistics/registered-provider-social-housing-stock-and-rents-in-england-2023-to-2024",
]

# Known stable download fallback — updated monthly by RSH
# These are checked first before scraping the publication page
_RP_LIST_FALLBACK_SEARCH = (
    "https://www.gov.uk/government/publications?"
    "keywords=registered+providers+list&"
    "departments[]=regulator-of-social-housing"
)

# RSH provider types
_PRP = "Private Registered Provider"
_LRP = "Local Authority Registered Provider"


class RSHRegisteredProvidersScraper:
    """Scraper for RSH Registered Providers list and SDR stock data.

    Uses plain httpx (not BaseScraper) since these are one-off bulk
    downloads rather than paginated scrapes.
    """

    def __init__(self, timeout: float = 60.0) -> None:
        self._timeout = timeout

    # ------------------------------------------------------------------
    # Registered Providers list
    # ------------------------------------------------------------------

    async def fetch_registered_providers(self) -> list[dict[str, Any]]:
        """Download and parse the RSH Registered Providers list.

        Tries to locate the latest CSV/XLSX download link from the GOV.UK
        publication page, then downloads and parses it.

        Returns
        -------
        list[dict]
            One dict per provider with keys:
            ``registration_number``, ``name``, ``provider_type``,
            ``status``, ``designation``, ``address``, ``stock_units``.
        """
        async with httpx.AsyncClient(
            timeout=httpx.Timeout(self._timeout),
            headers={"User-Agent": "Mozilla/5.0 (compatible; UKOpsBD/1.0)"},
            follow_redirects=True,
        ) as client:
            # Try the main publication page first
            download_url = await self._find_rp_download_url(client)

            if not download_url:
                logger.warning("rsh_rp_list_no_download_url_found")
                return []

            logger.info("rsh_rp_list_downloading", url=download_url)
            resp = await client.get(download_url)
            resp.raise_for_status()

            content_type = resp.headers.get("content-type", "").lower()
            if "spreadsheetml" in content_type or download_url.endswith((".xlsx", ".xls")):
                return self._parse_rp_excel(resp.content)
            else:
                return self._parse_rp_csv(resp.text)

    async def _find_rp_download_url(self, client: httpx.AsyncClient) -> str | None:
        """Scrape the GOV.UK publication page to find the latest RP list download."""
        for page_url in [_RP_LIST_URL, _RP_PUBLICATION_URL]:
            try:
                resp = await client.get(page_url)
                if resp.status_code != 200:
                    continue
                url = self._extract_download_link(resp.text, [".csv", ".xlsx", ".xls"])
                if url:
                    logger.info("rsh_rp_list_found_url", source_page=page_url, url=url)
                    return url
            except Exception as exc:
                logger.warning("rsh_rp_list_page_fetch_failed", url=page_url, error=str(exc))

        # Try a direct GOV.UK search for the publication
        try:
            resp = await client.get(
                "https://www.gov.uk/government/publications",
                params={
                    "keywords": "registered providers list social housing",
                    "departments[]": "regulator-of-social-housing",
                },
            )
            if resp.status_code == 200:
                # Find publication links and check the first result
                soup = BeautifulSoup(resp.text, "html.parser")
                for link in soup.select("a[href]"):
                    href = str(link.get("href", ""))
                    if "registered-providers" in href and "/publications/" in href:
                        full_url = href if href.startswith("http") else _GOV_BASE + href
                        pub_resp = await client.get(full_url)
                        if pub_resp.status_code == 200:
                            url = self._extract_download_link(
                                pub_resp.text, [".csv", ".xlsx", ".xls"]
                            )
                            if url:
                                return url
        except Exception as exc:
            logger.warning("rsh_rp_list_search_failed", error=str(exc))

        return None

    @staticmethod
    def _extract_download_link(html: str, extensions: list[str]) -> str | None:
        """Find the first download link with one of the given extensions."""
        soup = BeautifulSoup(html, "html.parser")
        for link in soup.find_all("a", href=True):
            href = str(link.get("href", ""))
            href_lower = href.lower()
            if any(href_lower.endswith(ext) for ext in extensions):
                if href.startswith("http"):
                    return href
                if href.startswith("/"):
                    return _GOV_BASE + href
        # Also look for asset links (GOV.UK CDN)
        for link in soup.find_all("a", href=True):
            href = str(link.get("href", ""))
            if "assets.publishing" in href:
                href_lower = href.lower()
                if any(ext in href_lower for ext in extensions):
                    return href
        return None

    def _parse_rp_excel(self, content: bytes) -> list[dict[str, Any]]:
        """Parse an XLSX registered providers file."""
        try:
            import openpyxl
        except ImportError:
            logger.error("rsh_rp_openpyxl_missing")
            return []

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return []

        # Find header row (first row with 'name' or 'registration' in it)
        header_idx = 0
        for i, row in enumerate(rows[:5]):
            row_text = " ".join(str(c).lower() for c in row if c)
            if any(kw in row_text for kw in ["registration", "provider name", "name"]):
                header_idx = i
                break

        headers = [str(c).lower().strip() if c else "" for c in rows[header_idx]]
        return self._map_rp_rows(headers, rows[header_idx + 1:])

    def _parse_rp_csv(self, text: str) -> list[dict[str, Any]]:
        """Parse a CSV registered providers file."""
        import csv as csv_mod
        reader = csv_mod.DictReader(io.StringIO(text))
        headers = [h.lower().strip() for h in (reader.fieldnames or [])]
        rows = list(reader)
        return self._map_rp_rows_from_dicts(rows)

    def _map_rp_rows_from_dicts(self, rows: list[dict]) -> list[dict[str, Any]]:
        """Map CSV DictReader rows to canonical provider dicts."""
        results = []
        for row in rows:
            # Normalise keys
            norm = {k.lower().strip(): v for k, v in row.items()}
            provider = self._extract_provider_fields(norm)
            if provider:
                results.append(provider)
        logger.info("rsh_rp_csv_parsed", count=len(results))
        return results

    def _map_rp_rows(
        self, headers: list[str], rows: list[tuple]
    ) -> list[dict[str, Any]]:
        """Map Excel rows (with header list) to canonical provider dicts."""
        results = []
        for row in rows:
            if not any(row):
                continue
            norm = {headers[i]: (str(row[i]).strip() if row[i] is not None else "")
                    for i in range(min(len(headers), len(row)))}
            provider = self._extract_provider_fields(norm)
            if provider:
                results.append(provider)
        logger.info("rsh_rp_excel_parsed", count=len(results))
        return results

    @staticmethod
    def _extract_provider_fields(norm: dict[str, str]) -> dict[str, Any] | None:
        """Extract canonical fields from a normalised row dict.

        Handles multiple possible column naming conventions used across
        different years of the RSH publication.
        """
        # Registration number — try various column names
        reg_num = (
            norm.get("registration number")
            or norm.get("reg no")
            or norm.get("registration no")
            or norm.get("reg number")
            or norm.get("rp number")
            or ""
        ).strip()

        # Provider name
        name = (
            norm.get("name")
            or norm.get("provider name")
            or norm.get("registered provider")
            or norm.get("organisation name")
            or ""
        ).strip()

        if not name:
            return None

        # Provider type: PRP / LRP
        ptype_raw = (
            norm.get("type")
            or norm.get("provider type")
            or norm.get("organisation type")
            or ""
        ).strip().upper()
        if "LRP" in ptype_raw or "LOCAL" in ptype_raw or "AUTHORITY" in ptype_raw:
            provider_type = "LRP"
        elif "PRP" in ptype_raw or "PRIVATE" in ptype_raw:
            provider_type = "PRP"
        else:
            provider_type = "PRP"  # Default — most are PRPs

        # Status: registered / de-registered
        status = (
            norm.get("status")
            or norm.get("registration status")
            or "Registered"
        ).strip()

        # Designation: housing association / other
        designation = (
            norm.get("designation")
            or norm.get("rp designation")
            or ""
        ).strip()

        # Address fields
        address_parts = []
        for col in ["address 1", "address1", "street", "town", "city", "county", "postcode"]:
            val = norm.get(col, "").strip()
            if val:
                address_parts.append(val)
        address = ", ".join(address_parts) if address_parts else (
            norm.get("address", "") or norm.get("registered office", "")
        ).strip()

        # Stock size (total units)
        units_raw = (
            norm.get("total homes")
            or norm.get("stock")
            or norm.get("units")
            or norm.get("total stock")
            or norm.get("social homes")
            or ""
        ).strip()
        stock_units: int | None = None
        if units_raw:
            try:
                stock_units = int(re.sub(r"[^\d]", "", units_raw))
            except ValueError:
                pass

        return {
            "registration_number": reg_num,
            "name": name,
            "provider_type": provider_type,
            "status": status,
            "designation": designation,
            "address": address,
            "stock_units": stock_units,
        }

    # ------------------------------------------------------------------
    # Statistical Data Return (SDR) — managed vs owned stock
    # ------------------------------------------------------------------

    async def fetch_sdr_stock(self) -> list[dict[str, Any]]:
        """Download and parse the most recent SDR stock data.

        Returns
        -------
        list[dict]
            One dict per provider with keys:
            ``name``, ``registration_number``,
            ``units_owned``, ``units_managed``, ``units_managed_for_others``,
            ``year``.
        """
        async with httpx.AsyncClient(
            timeout=httpx.Timeout(self._timeout),
            headers={"User-Agent": "Mozilla/5.0 (compatible; UKOpsBD/1.0)"},
            follow_redirects=True,
        ) as client:
            download_url = await self._find_sdr_download_url(client)
            if not download_url:
                logger.warning("rsh_sdr_no_download_url_found")
                return []

            logger.info("rsh_sdr_downloading", url=download_url)
            resp = await client.get(download_url)
            resp.raise_for_status()
            return self._parse_sdr_excel(resp.content)

    async def _find_sdr_download_url(self, client: httpx.AsyncClient) -> str | None:
        """Locate the latest SDR stock data download.

        Tries a list of known publication URLs (most recent first), then
        falls back to a GOV.UK search for the stock/rents publication.
        """
        # Try known publication URLs first
        for pub_url in _SDR_PUBLICATION_URLS:
            try:
                resp = await client.get(pub_url)
                if resp.status_code == 200:
                    url = self._extract_sdr_xlsx(resp.text)
                    if url:
                        logger.info("rsh_sdr_found_url", source=pub_url, url=url)
                        return url
            except Exception as exc:
                logger.debug("rsh_sdr_pub_url_failed", url=pub_url, error=str(exc))

        # Fallback: GOV.UK search
        try:
            resp = await client.get(
                "https://www.gov.uk/search/all",
                params={
                    "keywords": "registered provider social housing stock rents",
                    "organisations[]": "regulator-of-social-housing",
                    "content_purpose_supergroup[]": "research_and_statistics",
                    "order": "updated-newest",
                },
            )
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, "html.parser")
                for link in soup.select("a[href]"):
                    href = str(link.get("href", ""))
                    text = link.get_text(strip=True).lower()
                    if (
                        ("stock" in text or "rents" in text)
                        and "registered-provider" in href
                        and "/statistics/" in href
                    ):
                        pub_url = href if href.startswith("http") else _GOV_BASE + href
                        pub_resp = await client.get(pub_url)
                        if pub_resp.status_code == 200:
                            url = self._extract_sdr_xlsx(pub_resp.text)
                            if url:
                                return url
        except Exception as exc:
            logger.warning("rsh_sdr_search_failed", error=str(exc))

        return None

    @staticmethod
    def _extract_sdr_xlsx(html: str) -> str | None:
        """Find the combined/data XLSX download link in an SDR publication page."""
        soup = BeautifulSoup(html, "html.parser")
        # Prefer the "combined tool" or data XLSX (not the briefing note PDF)
        for link in soup.find_all("a", href=True):
            href = str(link.get("href", ""))
            if "assets.publishing" in href and href.lower().endswith(".xlsx"):
                href_lower = href.lower()
                # Skip if it looks like a methodology or technical notes file
                if any(skip in href_lower for skip in ["brief", "method", "technical", "notes", "quality"]):
                    continue
                return href
        # Fallback: any xlsx
        for link in soup.find_all("a", href=True):
            href = str(link.get("href", ""))
            if "assets.publishing" in href and href.lower().endswith(".xlsx"):
                return href
        return None

    def _parse_sdr_excel(self, content: bytes) -> list[dict[str, Any]]:
        """Parse the RSH stock-and-rents Excel file to extract per-RP stock figures.

        The 2025 publication uses a 'Combined Tool' workbook.  The relevant
        sheets are:

        * ``totals_&_RP_counts`` — RP_Code, RP_Name, Total Social Stock
        * ``STOCK_BY_LA``        — RP_Name, RP_Code, RP_Type, per-LA stock

        We aggregate STOCK_BY_LA to get total stock per RP (owned stock,
        since this publication reports owned/managed stock by LA).
        """
        try:
            import openpyxl
        except ImportError:
            logger.error("rsh_sdr_openpyxl_missing")
            return []

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        results: list[dict[str, Any]] = []

        # ── Try the totals sheet first (most direct) ──────────────────────────
        if "totals_&_RP_counts" in wb.sheetnames:
            results = self._parse_totals_sheet(wb["totals_&_RP_counts"])
            if results:
                logger.info("rsh_sdr_parsed_totals_sheet", count=len(results))
                wb.close()
                return results

        # ── Fall back to STOCK_BY_LA aggregation ──────────────────────────────
        if "STOCK_BY_LA" in wb.sheetnames:
            results = self._parse_stock_by_la_sheet(wb["STOCK_BY_LA"])
            if results:
                logger.info("rsh_sdr_parsed_stock_by_la", count=len(results))
                wb.close()
                return results

        # ── Generic fallback: scan all sheets ────────────────────────────────
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True, max_row=2000))
            if not rows:
                continue
            # Look for a header row containing "rp_name" or "provider name"
            for i, row in enumerate(rows[:15]):
                row_text = " ".join(str(c).lower() for c in row if c)
                if any(kw in row_text for kw in ["rp_name", "provider name", "registered provider"]):
                    headers = [
                        str(c).lower().strip().replace(" ", "_") if c else f"col_{j}"
                        for j, c in enumerate(row)
                    ]
                    for data_row in rows[i + 1:]:
                        if not any(data_row):
                            continue
                        norm = {
                            headers[j]: (str(data_row[j]).strip() if j < len(data_row) and data_row[j] is not None else "")
                            for j in range(len(headers))
                        }
                        name = norm.get("rp_name") or norm.get("provider_name") or norm.get("registered_provider") or ""
                        if not name:
                            continue
                        stock_raw = norm.get("total_social_stock") or norm.get("total_stock") or norm.get("stock") or ""
                        stock: int | None = None
                        if stock_raw:
                            try:
                                stock = int(re.sub(r"[^\d]", "", stock_raw))
                            except ValueError:
                                pass
                        results.append({
                            "name": name.strip(),
                            "registration_number": norm.get("rp_code") or norm.get("registration_number") or "",
                            "units_owned": stock,
                            "units_managed": None,
                            "units_managed_for_others": None,
                        })
                    break
            if results:
                break

        wb.close()
        logger.info("rsh_sdr_parsed", count=len(results))
        return results

    @staticmethod
    def _parse_totals_sheet(ws) -> list[dict[str, Any]]:
        """Parse the totals_&_RP_counts sheet: RP_Code, RP_Name, Total Social Stock."""
        results = []
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return results

        # Find header row
        header_idx = 0
        for i, row in enumerate(rows[:5]):
            row_text = " ".join(str(c).lower() for c in row if c)
            if "rp_code" in row_text or "rp_name" in row_text:
                header_idx = i
                break

        headers = [
            str(c).lower().strip().replace(" ", "_") if c else f"col_{j}"
            for j, c in enumerate(rows[header_idx])
        ]

        for row in rows[header_idx + 1:]:
            if not any(row):
                continue
            norm = {
                headers[j]: (str(row[j]).strip() if j < len(row) and row[j] is not None else "")
                for j in range(len(headers))
            }
            # The sheet has two sets of columns (left and right halves)
            # Use the first set: RP_Code, RP_Name, Total Social Stock
            name = norm.get("rp_name") or ""
            if not name or name.lower().startswith("rp_name"):
                continue
            reg = norm.get("rp_code") or ""
            stock_raw = norm.get("total_social_stock") or ""
            stock: int | None = None
            if stock_raw:
                try:
                    stock = int(re.sub(r"[^\d]", "", stock_raw))
                except ValueError:
                    pass
            results.append({
                "name": name,
                "registration_number": reg,
                "units_owned": stock,
                "units_managed": None,
                "units_managed_for_others": None,
            })
        return results

    @staticmethod
    def _parse_stock_by_la_sheet(ws) -> list[dict[str, Any]]:
        """Aggregate STOCK_BY_LA sheet to totals per RP."""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [
            str(c).lower().strip().replace(" ", "_") if c else f"col_{j}"
            for j, c in enumerate(rows[0])
        ]

        totals: dict[str, dict] = {}
        for row in rows[1:]:
            if not any(row):
                continue
            norm = {
                headers[j]: (str(row[j]).strip() if j < len(row) and row[j] is not None else "")
                for j in range(len(headers))
            }
            name = norm.get("rp_name") or ""
            if not name:
                continue
            reg = norm.get("rp_code") or ""
            stock_raw = norm.get("total_social_stock") or "0"
            try:
                stock = int(re.sub(r"[^\d]", "", stock_raw) or "0")
            except ValueError:
                stock = 0

            if name not in totals:
                totals[name] = {"name": name, "registration_number": reg, "units_owned": 0}
            totals[name]["units_owned"] = (totals[name]["units_owned"] or 0) + stock

        return [
            {**v, "units_managed": None, "units_managed_for_others": None}
            for v in totals.values()
        ]
